from datetime import datetime, timedelta
from PyQt5.QtWidgets import *
from PyQt5 import uic, QtWidgets
import sys
import openpyxl


class Date():
    def __init__(self, jour, mois, annee):
        self.jour = jour
        self.mois = mois
        self.annee = annee


class person():
    def __init__(self, cin, nom, prenom, age, adresse, nationalite, telephone, Date, decede):

        self.person = {
            'Cin': cin,
            'Nom': nom,
            'Prenom': prenom,
            'Age': age,
            'Adresse': adresse,
            'Nationalite': nationalite,
            'Telephone': telephone,
            'Date_infection': Date,
            'Decede': decede,
        }
        print(self.person)
        workbook = openpyxl.load_workbook('data.xlsx')
        worksheet = workbook.active
        worksheet['A1'] = 'Cin'
        worksheet['B1'] = 'Nom'
        worksheet['C1'] = 'Prenom'
        worksheet['D1'] = 'Age'
        worksheet['E1'] = 'Adresse'
        worksheet['F1'] = 'Nationalite'
        worksheet['G1'] = 'Telephone'
        worksheet['H1'] = 'Date_infection'
        worksheet['I1'] = 'Decede'

        worksheet.append([self.person['Cin'], self.person['Nom'], self.person['Prenom'], self.person['Age'], self.person['Adresse'],
                         self.person['Nationalite'], self.person['Telephone'], self.person['Date_infection'], self.person['Decede']])

        workbook.save('data.xlsx')


class maladie():
    def __init__(self, code, cinM, nomM, nbAnne):

        self.maladie = {
            'Code': code,
            'Cin': cinM,
            'Nom': nomM,
            'NBAnne': nbAnne,
        }
        print(self.maladie)
        workbook = openpyxl.load_workbook('maladie.xlsx')
        worksheet = workbook.active
        worksheet['A1'] = 'Code'
        worksheet['B1'] = 'Cin'
        worksheet['C1'] = 'Nom'
        worksheet['D1'] = 'NBAnne'

        worksheet.append([self.maladie['Code'], self.maladie['Cin'],
                         self.maladie['Nom'], self.maladie['NBAnne']])
        print("cin : ", self.maladie['Cin'], "nom : ",
              self.maladie['Nom'], "nbAnne : ", self.maladie['NBAnne'])

        workbook.save('maladie.xlsx')


class DeleteMaladie(QtWidgets.QTableWidget):
    def __init__(self):
        super(DeleteMaladie, self).__init__()
        uic.loadUi("Delete_M.ui", self)
        self.show()
        self.BtnDelete.clicked.connect(lambda: self.deleteEvent())

    def deleteEvent(self):
        var = self.codelineEdit.text()
        workbook = openpyxl.load_workbook('maladie.xlsx')
        worksheet = workbook.active
        Maladie_found = False
        for row in worksheet.iter_rows(min_row=2):
            if row[0].value == var:
                worksheet.delete_rows(row[0].row)
                workbook.save('maladie.xlsx')
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Information)
                msg.setWindowTitle("Success")
                msg.setText("Maladie deleted")
                msg.exec_()
                Maladie_found = True
                break
        if not Maladie_found:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Error")
            msg.setText("Maladie not found")
            msg.exec_()

        self.close()


class ModifyMaladie(QtWidgets.QDialog):
    def __init__(self, M_Object):
        super(ModifyMaladie, self).__init__()
        uic.loadUi("Modify_M.ui", self)
        self.show()
        self.BtnModify.clicked.connect(lambda: self.modifyEvent(M_Object))
        self.label.setText(M_Object)

    def modifyEvent(self, M_Object):
        var = self.cinDonneelineEdit.text()
        data = self.codelineEdit.text()
        workbook = openpyxl.load_workbook('maladie.xlsx')
        worksheet = workbook.active
        M_found = False
        self.label.setText(M_Object)
        workbook_data = openpyxl.load_workbook('data.xlsx')
        worksheet_data = workbook_data.active
        if M_Object == "NBAnne":
            for row in worksheet.iter_rows(min_row=2) :
                
                if row[1].value == var :
                   
                    M_found = True
                    workbook.save('maladie.xlsx')
                    msg = QMessageBox()
                    msg.setIcon(QMessageBox.Information)
                    msg.setWindowTitle("Success")
                    msg.setText("Maladie modified")
                    msg.exec_()
                    break
                
        elif M_Object == "nbDeces":
            for row_data in worksheet_data.iter_rows(min_row=2):
                if row_data[0].value == var:
                    row_data[8].value = data
                    workbook_data.save('data.xlsx')
                    M_found = True
                    break
                else:
                    msg = QMessageBox()
                    msg.setIcon(QMessageBox.Information)
                    msg.setWindowTitle("Error")
                    msg.setText("CIN not found")
                    msg.exec_()
                    break

        if not M_found:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Error")
            msg.setText("Error 502 found")
            msg.exec_()

        self.close()


class Diplay(QtWidgets.QTableWidget):
    def __init__(self):
        super().__init__(0, 9)
        self.initUI()

    def initUI(self):
        workbook = openpyxl.load_workbook('data.xlsx')
        worksheet = workbook.active

        self.setWindowTitle("Display")
        self.setGeometry(100, 100, 500, 500)
        self.verticalHeader().setVisible(False)

        self.setHorizontalHeaderLabels(
            ["Cin", "Nom", "Prenom", "Age", "Adresse", "Nationalite", "Telephone", "Date_infection", "Decede"])

        for i, row in enumerate(worksheet.iter_rows(min_row=2)):
            person = {
                'Cin': row[0].value,
                'Nom': row[1].value,
                'Prenom': row[2].value,
                'Age': row[3].value,
                'Adresse': row[4].value,
                'Nationalite': row[5].value,
                'Telephone': row[6].value,
                'Date_infection': row[7].value,
                'Decede': row[8].value,
            }
            self.insertRow(i)
            self.setItem(i, 0, QtWidgets.QTableWidgetItem(str(person['Cin'])))
            self.setItem(i, 1, QtWidgets.QTableWidgetItem(str(person['Nom'])))
            self.setItem(i, 2, QtWidgets.QTableWidgetItem(
                str(person['Prenom'])))
            self.setItem(i, 3, QtWidgets.QTableWidgetItem(str(person['Age'])))
            self.setItem(i, 4, QtWidgets.QTableWidgetItem(
                str(person['Adresse'])))
            self.setItem(i, 5, QtWidgets.QTableWidgetItem(
                str(person['Nationalite'])))
            self.setItem(i, 6, QtWidgets.QTableWidgetItem(
                str(person['Telephone'])))
            self.setItem(i, 7, QtWidgets.QTableWidgetItem(
                str(person['Date_infection'])))
            self.setItem(i, 8, QtWidgets.QTableWidgetItem(
                str(person['Decede'])))

        workbook.close()

        self.setSortingEnabled(True)
        self.horizontalHeader().setStretchLastSection(True)
        self.show()


class Search_Display(QtWidgets.QDialog):
    def __init__(self, S_Object):
        super(Search_Display, self).__init__()
        self.setWindowTitle("Search Display")
        self.tableWidget = QtWidgets.QTableWidget(self)
        self.tableWidget.setColumnCount(9)
        self.tableWidget.setHorizontalHeaderLabels(
            ["Cin", "Nom", "Prenom", "Age", "Adresse", "Nationalite", "Telephone", "Date_infection", "Decede"])
        self.layout = QtWidgets.QVBoxLayout(self)
        self.layout.addWidget(self.tableWidget)
        self.show()
        self.label = QtWidgets.QLabel(S_Object, self)
        self.layout.addWidget(self.label)
        self.lineEdit = QtWidgets.QLineEdit(self)
        self.layout.addWidget(self.lineEdit)
        self.BtnSearch = QtWidgets.QPushButton('Search', self)
        self.layout.addWidget(self.BtnSearch)
        self.BtnSearch.clicked.connect(lambda: self.searchEvent(S_Object))

    def searchEvent(self, S_Object):
        var = self.lineEdit.text()  # val passÃ© en parametre
        workbook = openpyxl.load_workbook('data.xlsx')
        worksheet = workbook.active
        row_count = worksheet.max_row - 1
        self.tableWidget.setRowCount(row_count)
        row_index = 0
        for row in worksheet.iter_rows(min_row=2):
            if S_Object == "Telephone":
                if row[6].value == var:
                    person = [row[0].value, row[1].value, row[2].value, row[3].value,
                              row[4].value, row[5].value, row[6].value, row[7].value, row[8].value]
                    for col_index, value in enumerate(person):
                        self.tableWidget.setItem(
                            row_index, col_index, QtWidgets.QTableWidgetItem(str(value)))
                    row_index += 1
            elif S_Object == "Cin":
                if row[0].value == var:
                    person = [row[0].value, row[1].value, row[2].value, row[3].value,
                              row[4].value, row[5].value, row[6].value, row[7].value, row[8].value]
                    for col_index, value in enumerate(person):
                        self.tableWidget.setItem(
                            row_index, col_index, QtWidgets.QTableWidgetItem(str(value)))
                    row_index += 1
            elif S_Object == "Nationalite":
                if row[5].value == var:
                    person = [row[0].value, row[1].value, row[2].value, row[3].value,
                              row[4].value, row[5].value, row[6].value, row[7].value, row[8].value]
                    for col_index, value in enumerate(person):
                        self.tableWidget.setItem(
                            row_index, col_index, QtWidgets.QTableWidgetItem(str(value)))
                    row_index += 1
            elif S_Object == "D":
                if row[8].value == "1":
                    person = [row[0].value, row[1].value, row[2].value, row[3].value,
                              row[4].value, row[5].value, row[6].value, row[7].value, row[8].value]
                    for col_index, value in enumerate(person):
                        self.tableWidget.setItem(
                            row_index, col_index, QtWidgets.QTableWidgetItem(str(value)))
                    row_index += 1
            elif S_Object == "ND":
                if row[8].value == "0":
                    person = [row[0].value, row[1].value, row[2].value, row[3].value,
                              row[4].value, row[5].value, row[6].value, row[7].value, row[8].value]
                    for col_index, value in enumerate(person):
                        self.tableWidget.setItem(
                            row_index, col_index, QtWidgets.QTableWidgetItem(str(value)))
                    row_index += 1
        workbook.close()


class DisplayMaladie(QtWidgets.QTableWidget):
    def __init__(self, D_Object):

        i = 0
        if D_Object == "ALL":
            super().__init__(0, 4)
            self.setWindowTitle("Afficher les Maladies")
            self.setGeometry(100, 100, 500, 500)
            self.verticalHeader().setVisible(False)
            self.setHorizontalHeaderLabels(["Code", "Cin", "Nom", "NBAnne"])
            workbook = openpyxl.load_workbook('maladie.xlsx')
            worksheet = workbook.active
            for j, row in enumerate(worksheet.iter_rows(min_row=2)):
                maladie = {
                    'Code': row[0].value,
                    'Cin': row[1].value,
                    'Nom': row[2].value,
                    'NBAnne': row[3].value,

                }
                self.insertRow(j)
                self.setItem(j, 0, QtWidgets.QTableWidgetItem(
                    str(maladie['Code'])))
                self.setItem(j, 1, QtWidgets.QTableWidgetItem(
                    str(maladie['Cin'])))
                self.setItem(j, 2, QtWidgets.QTableWidgetItem(
                    str(maladie['Nom'])))
                self.setItem(j, 3, QtWidgets.QTableWidgetItem(
                    str(maladie['NBAnne'])))

            workbook.close()
        elif D_Object == "Pourcentage":
            super().__init__(0, 2)
            self.setWindowTitle("Afficher les Maladies")
            self.setGeometry(100, 100, 500, 500)
            self.verticalHeader().setVisible(False)
            self.setHorizontalHeaderLabels(["Nom", "Pourcentage"])
            workbook = openpyxl.load_workbook('maladie.xlsx')
            worksheet = workbook.active
            for row in worksheet.iter_rows(min_row=2):
                count = 0
                var_M = row[2].value
                for row_M in worksheet.iter_rows(min_row=2):
                    if var_M == row_M[2].value:
                        count += 1
                self.insertRow(i)
                self.setItem(i, 0, QtWidgets.QTableWidgetItem(str(var_M)))
                self.setItem(i, 1, QtWidgets.QTableWidgetItem(
                    str(str((count / (worksheet.max_row - 1)) * 100)+"%")))

                i += 1
            workbook.close()
        elif D_Object == "All_P":
            super().__init__(0, 8)

            self.setWindowTitle("Afficher les Maladies")
            self.setGeometry(100, 100, 500, 500)
            self.verticalHeader().setVisible(False)
            self.setHorizontalHeaderLabels(
                ["Cin", "Nom", "Prenom", "Age", "Adresse", "Nationalite", "Telephone", "Maladie(s)"])
            workbook_data = openpyxl.load_workbook('data.xlsx')
            worksheet_data = workbook_data.active
            array = []
            mal = []
            i = 0
            for row_data in worksheet_data.iter_rows(min_row=2):
                data = []
                workbook = openpyxl.load_workbook('maladie.xlsx')
                worksheet = workbook.active

                Cin = row_data[0].value
                Nom = row_data[1].value
                Prenom = row_data[2].value
                Age = row_data[3].value
                Adresse = row_data[4].value
                Nationalite = row_data[5].value
                Telephone = row_data[6].value
                Date_infection = row_data[7].value

                pog = {
                    'Cin': Cin,
                    'Nom': Nom,
                    'Prenom': Prenom,
                    'Age': Age,
                    'Adresse': Adresse,
                    'Nationalite': Nationalite,
                    'Telephone': Telephone,
                    'Date_infection': Date_infection,

                }

                array.append(pog)
                for row in worksheet.iter_rows(min_row=2):
                    data.append(row[1].value)
                if Cin in data:
                    for row in worksheet.iter_rows(min_row=2):
                        if row[1].value == Cin:
                            mal.append(row[2].value)
                    workbook.close()

                    self.insertRow(i)
                    self.setItem(
                        i, 0, QtWidgets.QTableWidgetItem(str(pog['Cin'])))
                    self.setItem(
                        i, 1, QtWidgets.QTableWidgetItem(str(pog['Nom'])))
                    self.setItem(i, 2, QtWidgets.QTableWidgetItem(
                        str(pog['Prenom'])))
                    self.setItem(
                        i, 3, QtWidgets.QTableWidgetItem(str(pog['Age'])))
                    self.setItem(i, 4, QtWidgets.QTableWidgetItem(
                        str(pog['Adresse'])))
                    self.setItem(i, 5, QtWidgets.QTableWidgetItem(
                        str(pog['Nationalite'])))
                    self.setItem(i, 6, QtWidgets.QTableWidgetItem(
                        str(pog['Telephone'])))
                    self.setItem(i, 7, QtWidgets.QTableWidgetItem(
                        str(pog['Date_infection'])))
                    self.setItem(i, 1, QtWidgets.QTableWidgetItem(str(mal)))

                    i += 1
            workbook_data.close()

        self.setSortingEnabled(True)
        self.horizontalHeader().setStretchLastSection(True)
        self.show()


class Search_Display_Maladie(QtWidgets.QDialog):
    def __init__(self, S_Object):
        super(Search_Display_Maladie, self).__init__()
        self.setWindowTitle("Affichage")
        self.tableWidget = QtWidgets.QTableWidget(self)
        self.tableWidget.setColumnCount(5)
        self.tableWidget.setHorizontalHeaderLabels(
            ["Code", "Cin", "Nom", "NBAnne"])
        self.layout = QtWidgets.QVBoxLayout(self)
        self.layout.addWidget(self.tableWidget)
        self.show()
        self.label = QtWidgets.QLabel(S_Object, self)
        self.layout.addWidget(self.label)
        self.lineEdit = QtWidgets.QLineEdit(self)
        self.layout.addWidget(self.lineEdit)
        self.BtnSearch = QtWidgets.QPushButton('Search', self)
        self.layout.addWidget(self.BtnSearch)
        self.BtnSearch.clicked.connect(lambda: self.searchEvent(S_Object))

    def searchEvent(self, S_Object):
        var = self.lineEdit.text()  # val passÃ© en parametre
        workbook = openpyxl.load_workbook('maladie.xlsx')
        worksheet = workbook.active
        row_count = worksheet.max_row - 1
        self.tableWidget.setRowCount(row_count)
        row_index = 0
        col_index = 0

        for row in worksheet.iter_rows(min_row=2):

            if S_Object == "Nom":
                if row[2].value == var:
                    maladie = [row[0].value, row[1].value,
                               row[2].value, row[3].value]
                    for col_index, value in enumerate(maladie):
                        self.tableWidget.setItem(
                            row_index, col_index, QtWidgets.QTableWidgetItem(str(value)))
                    row_index += 1
            elif S_Object == "Cin":
                if row[1].value == var:
                    maladie = [row[0].value, row[1].value,
                               row[2].value, row[3].value]
                    for col_index, value in enumerate(maladie):
                        self.tableWidget.setItem(
                            row_index, col_index, QtWidgets.QTableWidgetItem(str(value)))
                    row_index += 1
        workbook.close()


class ModifyWindow(QtWidgets.QDialog):
    def __init__(self, M_Object):
        super(ModifyWindow, self).__init__()
        uic.loadUi("Modify.ui", self)
        self.show()
        self.BtnModify.clicked.connect(lambda: self.modifyEvent(M_Object))
        self.label.setText(M_Object)

    def modifyEvent(self, M_Object):
        var = self.M_LineEdit.text()
        cin = self.M_LineEdit_2.text()
        workbook = openpyxl.load_workbook('data.xlsx')
        worksheet = workbook.active
        person_found = False

        for row in worksheet.iter_rows(min_row=2):
            if M_Object == "Adresse":
                if row[0].value == cin:
                    row[4].value = var
                    workbook.save('data.xlsx')
                    msg = QMessageBox()
                    msg.setIcon(QMessageBox.Information)
                    msg.setWindowTitle("Success")
                    msg.setText("Person modified")
                    msg.exec_()
                    person_found = True
                    break
            elif M_Object == "Telephone":
                if row[0].value == cin:
                    row[6].value = var
                    workbook.save('data.xlsx')
                    msg = QMessageBox()
                    msg.setIcon(QMessageBox.Information)
                    msg.setWindowTitle("Success")
                    msg.setText("Person modified")
                    msg.exec_()
                    person_found = True
                    break

        if not person_found:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Error")
            msg.setText("Person not found")
            msg.exec_()

        self.close()


class DeleteWindow(QtWidgets.QDialog):
    def __init__(self, D_Object):
        super(DeleteWindow, self).__init__()
        uic.loadUi("Delete.ui", self)
        self.show()
        self.BtnDelete.clicked.connect(lambda: self.deleteEvent(D_Object))
        self.label.setText(D_Object)

    def deleteEvent(self, D_Object):
        var = self.cinLineEdit.text()
        workbook = openpyxl.load_workbook('data.xlsx')
        worksheet = workbook.active
        person_found = False
        for row in worksheet.iter_rows(min_row=2):
            if D_Object == "Cin":
                if row[0].value == var:
                    worksheet.delete_rows(row[0].row)
                    workbook.save('data.xlsx')
                    msg = QMessageBox()
                    msg.setIcon(QMessageBox.Information)
                    msg.setWindowTitle("Success")
                    msg.setText("Person deleted")
                    msg.exec_()
                    person_found = True
                    break
            elif D_Object == "Nationalite":
                if row[5].value == var:
                    worksheet.delete_rows(row[0].row)
                    workbook.save('data.xlsx')
                    msg = QMessageBox()
                    msg.setIcon(QMessageBox.Information)
                    msg.setWindowTitle("Success")
                    msg.setText("Person deleted")
                    msg.exec_()
                    person_found = True
                    break
            elif D_Object == "Telephone":
                if row[6].value == var:
                    worksheet.delete_rows(row[0].row)
                    workbook.save('data.xlsx')
                    msg = QMessageBox()
                    msg.setIcon(QMessageBox.Information)
                    msg.setWindowTitle("Success")
                    msg.setText("Person deleted")
                    msg.exec_()
                    person_found = True
                    break
        if not person_found:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Error")
            msg.setText("Person not found")
            msg.exec_()

        self.close()


class Affiche(QtWidgets.QDialog):
    def __init__(self, S_Object):
        super(Affiche, self).__init__()
        uic.loadUi("Calcul et Affichage.ui", self)
        self.show()
        self.tableWidget.setHorizontalHeaderLabels(
            ["Cin", "Nom", "Prenom", "Age", "Adresse", "Nationalite", "Telephone", "Maladie(s)"])
        self.tableWidget.setColumnCount(9)
        self.tableWidget.setRowCount(10)
        self.tableWidget.setItem(0, 0, QtWidgets.QTableWidgetItem("Cin"))
        self.BtnSearch.clicked.connect(lambda: self.AfficheEvent(S_Object))
        self.label.setText(S_Object)

    def AfficheEvent(self, S_Object):
        var = self.lineEdit.text()
        workbook = openpyxl.load_workbook('data.xlsx')
        worksheet = workbook.active
        workbook_m = openpyxl.load_workbook('maladie.xlsx')
        worksheet_m = workbook_m.active
        current_date = datetime.strptime('2022-04-02', '%Y-%m-%d').date()
        quarantine_filtered = []
        array = []
        show = []
        row_index = 0
        i = 0
        x = 0
        risk = 0
        if S_Object == "Nationalite":
            for row in worksheet.iter_rows(min_row=2):
                if row[5].value == var:
                    person = [row[0].value, row[1].value, row[2].value, row[3].value,
                              row[4].value, row[5].value, row[6].value, row[7].value, row[8].value]
                    for col_index, value in enumerate(person):
                        self.tableWidget.setHorizontalHeaderLabels(
                            ["Cin", "Nom", "Prenom", "Age", "Adresse", "Nationalite", "Telephone", "Maladie(s)"])
                        self.tableWidget.setColumnCount(9)
                        self.tableWidget.setItem(
                            row_index, col_index, QTableWidgetItem(str(value)))
                    row_index += 1
            workbook.close()
            workbook.close()
        elif S_Object == "Cin":
            for row in worksheet.iter_rows(min_row=2):
                self.tableWidget.setHorizontalHeaderLabels(
                    ["Cin", "Nom", "Prenom", "Age", "Adresse", "Nationalite", "Telephone", "Maladie(s)"])
                self.tableWidget.setColumnCount(9)
                data = []
                mal = []
                Cin = row[0].value
                Nom = row[1].value
                Prenom = row[2].value
                Age = row[3].value
                Adresse = row[4].value
                Nationalite = row[5].value
                Telephone = row[6].value
                Date_infection = row[7].value

                pog = {
                    'Cin': Cin,
                    'Nom': Nom,
                    'Prenom': Prenom,
                    'Age': Age,
                    'Adresse': Adresse,
                    'Nationalite': Nationalite,
                    'Telephone': Telephone,
                    'Date_infection': Date_infection,
                }

                array.append(pog)
                for row_m in worksheet_m.iter_rows(min_row=2):
                    data.append(row_m[1].value)

                if Cin in data:
                    for p in (pog['Date_infection'].strftime('%Y-%m-%d')):

                        p_date = datetime.strptime(
                            pog['Date_infection'].strftime('%Y-%m-%d'), '%Y-%m-%d').date()
                        days_since_infection = (current_date - p_date).days
                        if days_since_infection <= 14:
                            quarantine_filtered.append(days_since_infection)

                    self.tableWidget.insertRow(i)
                    self.tableWidget.setItem(
                        i, 0, QTableWidgetItem(str(pog['Cin'])))
                    self.tableWidget.setItem(
                        i, 1, QTableWidgetItem(str(pog['Nom'])))
                    self.tableWidget.setItem(
                        i, 2, QTableWidgetItem(str(pog['Prenom'])))
                    self.tableWidget.setItem(
                        i, 3, QTableWidgetItem(str(pog['Age'])))
                    self.tableWidget.setItem(
                        i, 4, QTableWidgetItem(str(pog['Adresse'])))
                    self.tableWidget.setItem(
                        i, 5, QTableWidgetItem(str(pog['Nationalite'])))
                    self.tableWidget.setItem(
                        i, 6, QTableWidgetItem(str(pog['Telephone'])))
                    self.tableWidget.setItem(
                        i, 7, QTableWidgetItem(str(pog['Date_infection'])))
                    self.tableWidget.setItem(i, 8, QTableWidgetItem(str(mal)))
                    i += 1
                    self.tableWidget.setRowCount(i)
                    self.tableWidget.resizeColumnsToContents()
                    self.tableWidget.resizeRowsToContents()
                    self.tableWidget.show()
                    workbook.close()
                    workbook_m.close()
        elif S_Object == "D":
            self.tableWidget.setColumnCount(2)
            for row in worksheet.iter_rows(min_row=2):
                if row[8].value == 1:
                    person = [row[0].value, row[1].value, row[2].value, row[3].value,
                              row[4].value, row[5].value, row[6].value, row[7].value, row[8].value]
                    for col_index, value in enumerate(person):
                        self.tableWidget.setItem(
                            row_index, col_index, QTableWidgetItem(str(value)))
                        row_index += 1

                    self.tableWidget.setHorizontalHeaderLabels(
                        ["Nom", "%Deces"])
                    count = 0
                    var_M = row[2].value
                    for row_M in worksheet.iter_rows(min_row=2):
                        if var_M == row_M[2].value:
                            count += 1
                    dict_M = {
                        var_M: (count / (worksheet.max_row - 1)) * 100

                    }
                    T = list(dict_M)
                    T = set(T)
                    x = 0
                    self.tableWidget.setRowCount(len(T))
                    print(dict_M)

                    for i in dict_M:
                        self.tableWidget.setItem(
                            x, 0, QTableWidgetItem(str(i)))
                        self.tableWidget.setItem(
                            x, 1, QTableWidgetItem(str(dict_M[i])))
                        x += 1

                    self.tableWidget.show()
                    workbook.close()
                    workbook_m.close()


class AffichePourcentage(QtWidgets.QTableWidget):
    def __init__(self):
        super().__init__(0, 3)
        self.initUI()

    def initUI(self):
        self.setWindowTitle("Affichage de Pourcentage")
        self.setGeometry(100, 100, 500, 500)
        self.verticalHeader().setVisible(False)
        self.setHorizontalHeaderLabels(["Nom", "Prenom", "PourcentageRisk"])
        self.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)
        self.setStyleSheet(
            "background-color: rgb(200, 194, 167); font-size: 12px;font-family: 'Segoe Script';")

        self.show()

        workbook_m = openpyxl.load_workbook('maladie.xlsx')
        worksheet_m = workbook_m.active
        workbook = openpyxl.load_workbook('data.xlsx')
        worksheet = workbook.active
        risk = 0
        i = 0
        for row in worksheet.iter_rows(min_row=2):
            for row_m in worksheet_m.iter_rows(min_row=2):
                if int(row[3].value) >= 70:
                    risk = 20
                elif 50 <= int(row[3].value) <= 70:
                    risk = 10
                elif row_m[2].value == "l'hypertension":
                    risk = 20
                elif row_m[2].value == "l'asthme":
                    risk = 20
                elif row_m[2].value == "le diabete":
                    risk = 15
                elif int(row[3].value) >= 70 and row_m[2].value == "l'hypertension" and row_m[2].value == "l'asthme" and row_m[2].value == "le diabete":
                    risk = 75
                elif int(row[3].value) >= 70 and row_m[2].value == "l'hypertension" and row_m[2].value == "l'asthme":
                    risk = 60
                elif int(row[3].value) >= 70 and row_m[2].value == "le diabete" and row_m[2].value == "l'asthme":
                    risk = 55
                elif int(row[3].value) >= 70 and row_m[2].value == "l'hypertension" and row_m[2].value == "le diabete":
                    risk = 55
                elif int(row[3].value) >= 70 and row_m[2].value == "l'hypertension":
                    risk = 40
                elif int(row[3].value) >= 70 and row_m[2].value == "l'asthme":
                    risk = 40
                elif int(row[3].value) >= 70 and row_m[2].value == "le diabete":
                    risk = 35

                elif 50 <= int(row[3].value) <= 70 and row_m[2].value == "l'hypertension" and row_m[2].value == "l'asthme" and row_m[2].value == "le diabete":
                    risk = 65
                elif 50 <= int(row[3].value) <= 70 and row_m[2].value == "l'hypertension" and row_m[2].value == "l'asthme":
                    risk = 50
                elif 50 <= int(row[3].value) <= 70 and row_m[2].value == "le diabete" and row_m[2].value == "l'asthme":
                    risk = 45
                elif 50 <= int(row[3].value) <= 70 and row_m[2].value == "l'hypertension" and row_m[2].value == "le diabete":
                    risk = 45
                elif 50 <= int(row[3].value) <= 70 and row_m[2].value == "l'hypertension":
                    risk = 30
                elif 50 <= int(row[3].value) <= 70 and row_m[2].value == "l'asthme":
                    risk = 30
                elif 50 <= int(row[3].value) <= 70 and row_m[2].value == "le diabete":
                    risk = 25

                elif row_m[2].value == "l'hypertension" and row_m[2].value == "l'asthme" and row_m[2].value == "le diabete":
                    risk = 55
                elif row_m[2].value == "l'hypertension" and row_m[2].value == "l'asthme":
                    risk = 40
                elif row_m[2].value == "le diabete" and row_m[2].value == "l'asthme":
                    risk = 35
                elif row_m[2].value == "l'hypertension" and row_m[2].value == "le diabete":
                    risk = 35
                self.insertRow(i)
                self.setItem(
                    i, 0, QtWidgets.QTableWidgetItem(str(row[1].value)))
                self.setItem(
                    i, 1, QtWidgets.QTableWidgetItem(str(row[2].value)))
                self.setItem(i, 2, QtWidgets.QTableWidgetItem(
                    str(str(risk)+"%")))
                i += 1
        workbook.close()
        workbook_m.close()
        self.setSortingEnabled(True)
        self.horizontalHeader().setStretchLastSection(True)
        self.show()


class MainWindow(QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        uic.loadUi("Main.ui", self)
        self.show()
        self.btn_ajouter.triggered.connect(self.add_person)
        self.action_Suppression_Personne_Donne.triggered.connect(
            self.delete_person)
        self.action_Suppression_Des_Personnes_d_une_natonalite_donnee.triggered.connect(
            self.delete_person1)
        self.action_Suppression_des_personnes_d_un_indicatif_donne_telephone.triggered.connect(
            self.delete_person2)
        self.actionTelephone_2.triggered.connect(self.modify_person)
        self.action_Adresse_2.triggered.connect(self.modify_person1)
        self.btn_all.triggered.connect(self.show_all)
        self.btn_search_tel.triggered.connect(self.search1)
        self.btn_indicatif.triggered.connect(self.search2)
        self.btn_nationalit.triggered.connect(self.search3)
        self.btn_D.triggered.connect(self.search4)
        self.btn_ND.triggered.connect(self.search5)
        self.btn_ajoutMaladie.triggered.connect(self.add_maladie)
        self.action_Supprimer_une_Maladie.triggered.connect(
            self.delete_maladie)
        self.action_Nombre_d_annees.triggered.connect(self.modify_maladie)
        self.action_Modifier_deces_de_0_a_1.triggered.connect(
            self.modify_maladie1)

        self.actionContenu_du_dictionnaire_maladies.triggered.connect(
            self.search8)
        self.actionRecherche_par_une_maladie.triggered.connect(self.search7)
        self.actionRecherche_maladies_d_une_personne.triggered.connect(
            self.search6)
        self.actionRecherche_le_pourcentage_de_chaque_maladie.triggered.connect(
            self.search9)
        self.actionRecherche_maladies_de_chaque_personne.triggered.connect(
            self.search10)

        self.action_afficher_par_nationalite.triggered.connect(self.affiche1)
        self.action_Personne_en_quarantaine.triggered.connect(self.affiche2)
        self.action_Personnes_Decedes.triggered.connect(self.affiche3)
        self.action_Personnes_A_risque.triggered.connect(self.affichage)

    def add_person(self):
        self.add_person = AddPerson()
        self.add_person.show()

    def delete_person(self):
        self.delete_person = DeleteWindow("Cin")
        self.delete_person.show()

    def delete_person1(self):
        self.delete_person = DeleteWindow("Nationalite")
        self.delete_person.show()

    def delete_person2(self):
        self.delete_person = DeleteWindow("Telephone")
        self.delete_person.show()

    def modify_person(self):
        self.modify_person = ModifyWindow("Telephone")
        self.modify_person.show()

    def modify_person1(self):
        self.modify_person = ModifyWindow("Adresse")
        self.modify_person.show()

    def show_all(self):
        self.show_all = Diplay()
        self.show_all.show()

    def search1(self):
        self.search = Search_Display("Telephone")
        self.search.show()

    def search2(self):
        self.search = Search_Display("Cin")
        self.search.show()

    def search3(self):
        self.search = Search_Display("Nationalite")
        self.search.show()

    def search4(self):
        self.search = Search_Display("D")
        self.search.show()

    def search5(self):
        self.search = Search_Display("ND")
        self.search.show()

    def add_maladie(self):
        self.add_maladie = AddMaladie()
        self.add_maladie.show()

    def delete_maladie(self):
        self.delete_maladie = DeleteMaladie()
        self.delete_maladie.show()

    def modify_maladie(self):
        self.modify_maladie = ModifyMaladie("NBAnne")
        self.modify_maladie.show()

    def modify_maladie1(self):
        self.modify_maladie = ModifyMaladie("nbDeces")
        self.modify_maladie.show()

    def show_all_maladies(self):
        self.show_all_maladies = Search_Display_Maladie()
        self.show_all_maladies.show()

    def search6(self):
        self.search = Search_Display_Maladie("Cin")
        self.search.show()

    def search7(self):
        self.search = Search_Display_Maladie("Nom")
        self.search.show()

    def search8(self):
        self.search = DisplayMaladie("ALL")
        self.search.show()

    def search9(self):
        self.search = DisplayMaladie("Pourcentage")
        self.search.show()

    def search10(self):
        self.search = DisplayMaladie("All_P")
        self.search.show()

    def affiche1(self):
        self.affiche = Affiche("Nationalite")
        self.affiche.show()

    def affiche2(self):
        self.affiche = Affiche("Cin")
        self.affiche.show()

    def affiche3(self):
        self.affiche = Affiche("D")
        self.affiche.show()

    def affichage(self):
        self.affichage = AffichePourcentage()
        self.affichage.show()


class AddMaladie(QtWidgets.QDialog):
    def __init__(self):
        super(AddMaladie, self).__init__()
        uic.loadUi("AddMaladie.ui", self)
        self.show()
        self.pushButton.clicked.connect(self.sendEvent0)
        self.CancelBtn.clicked.connect(self.close)

    def sendEvent0(self):
        if not self.validateFields():
            return
        code = self.codelineEdit.text()
        cinM = self.cinMlineEdit.text()
        nomM = self.nomMlineEdit.text()
        nbAnne = self.nbAnnelineEdit.text()
        maladie(code, cinM, nomM, nbAnne)

        self.close()

    def validateFields(self):
        if not self.codelineEdit.text().isdigit():
            msg = QMessageBox()
            msg.setWindowTitle("Error")
            msg.setText("Code must be a number")
            msg.exec_()
            return False
        if not self.cinMlineEdit.text().isdigit():
            msg = QMessageBox()
            msg.setWindowTitle("Error")
            msg.setText("Cin must be a number")
            msg.exec_()
            return False
        if len(self.nomMlineEdit.text()) == 1:
            msg = QMessageBox()
            msg.setWindowTitle("Error")
            msg.setText("invalid name")
            msg.exec_()
            return False
        return True


class AddPerson(QtWidgets.QDialog):
    def __init__(self):
        super(AddPerson, self).__init__()
        uic.loadUi("AddPerson.ui", self)
        self.show()
        self.SendBtn.clicked.connect(self.sendEvent)
        self.BtnBack.clicked.connect(self.close)

    def sendEvent(self):
        if not self.validateFields():
            return
        cin = self.cinLineEdit.text()
        nom = self.nomLineEdit.text()
        prenom = self.prenomLineEdit.text()
        age = self.ageLineEdit.text()
        adresse = self.adresseLineEdit.text()
        nationalite = self.nationaliteLineEdit.text()
        telephone = self.telephoneLineEdit.text()
        date_infection = self.date_infectionLineEdit.text()
        decede = self.decedeLineEdit.text()
        if self.check(cin, nom, prenom, age, adresse, nationalite, telephone, date_infection, decede) and not self.Search(cin):
            person(cin, nom, prenom, age, adresse, nationalite,
                   telephone, date_infection, decede)
        else:
            if self.Search(cin):
                msg = QMessageBox()
                msg.setWindowTitle("Error")
                msg.setText("Cin already exist")
                self.close()
            else:
                msg = QMessageBox()
                msg.setWindowTitle("Please fill all the fields")
        self.close()

    def check(self, cin, nom, prenom, age, adresse, nationalite, telephone, date_infection, decede):
        if cin == '' or nom == '' or prenom == '' or age == '' or adresse == '' or nationalite == '' or telephone == '' or date_infection == '' or decede == '':
            return False
        else:
            return True

    def Search(self, cin):
        workbook = openpyxl.load_workbook('data.xlsx')
        worksheet = workbook.active
        for row in worksheet.iter_rows(min_row=2):
            if row[0].value == cin:
                return True
        return False

    def validateFields(self):
        if not self.cinLineEdit.text().isdigit():
            msg = QMessageBox()
            msg.setWindowTitle("Error")
            msg.setText("Cin must be a number")
            msg.exec_()
            return False
        if not self.ageLineEdit.text().isdigit():
            msg = QMessageBox()
            msg.setWindowTitle("Error")
            msg.setText("Age must be a number")
            msg.exec_()
            return False
        if not int(self.ageLineEdit.text()) < 100:
            msg = QMessageBox()
            msg.setWindowTitle("Error")
            msg.setText("You are not a elf")
            msg.exec_()
            return False
        if not self.telephoneLineEdit.text().isdigit():
            msg = QMessageBox()
            msg.setWindowTitle("Error")
            msg.setText("Telephone must be a number")
            msg.exec_()
            return False
        
        if not len(self.date_infectionLineEdit.text()) == 10:
            msg = QMessageBox()
            msg.setWindowTitle("Error")
            msg.setText("Date must be a number")
            msg.exec_()
            return False
        if not self.decedeLineEdit.text().isdigit():
            msg = QMessageBox()
            msg.setWindowTitle("Error")
            msg.setText("Date must be a number")
            msg.exec_()
            return False
        '''if len(self.nomLineEdit.text()) == 1:
            msg = QMessageBox()
            msg.setWindowTitle("Error")
            msg.setText("invalid name")
            msg.exec_()
            return False'''
        return True


def main():
    app = QApplication([])
    window = MainWindow()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()